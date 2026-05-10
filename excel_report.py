"""Build and ingest load-test Excel reports.
"""

from __future__ import annotations

import io
import os
import json
from copy import copy
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import PatternFill, Font

import site_db
import model_loader


_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "templates", "load_test_template.xlsx"
)

_BLOCK_STRIDE = 6           # rows per block (4 content + 2 spacer)
_FIRST_BLOCK_ROW = 20       # first content row of block 1
_TEMPLATE_BLOCK_COUNT = 5   # blocks already laid out in the template

_RED_FILL = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
_BLUE_FILL = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")

_PHASE_COLS = {
    # phase: (label_col, mag_col, ang_col)
    "A": ("U",  "U",  "Y"),
    "B": ("AD", "AD", "AH"),
    "C": ("AM", "AM", "AQ"),
    "N": ("AV", "AV", "AZ"),
}

_CURRENT_KEYS = {
    "A": ("Phase A Current", "Phase A I-Angle"),
    "B": ("Phase B Current", "Phase B I-Angle"),
    "C": ("Phase C Current", "Phase C I-Angle"),
    "N": ("Neutral Current", "Neutral I-Angle"),
}
_VOLTAGE_KEYS = {
    "A": ("Phase A Voltage", "Phase A V-Angle"),
    "B": ("Phase B Voltage", "Phase B V-Angle"),
    "C": ("Phase C Voltage", "Phase C V-Angle"),
    "N": ("Neutral Voltage",  "Neutral V-Angle"),
}


# ── Report Generation ─────────────────────────────────────────────────────────

def build_load_test_report(db_path: str, test_id: str, use360: bool = True) -> bytes | None:
    data = site_db.get_test_report_data(db_path, test_id)
    if not data:
        return None

    topology = site_db.get_latest_topology(db_path) or {}
    topo_devices = {d.get("id"): d for d in topology.get("devices", [])}
    
    # Run simulation to get predicted values
    sim_devices = model_loader.load_substation_model(topology)
    sim_summaries = {did: dev.get_summary_dict() for did, dev in sim_devices.items()}

    wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    if "Template" in wb.sheetnames:
        del wb["Template"]
    ws = wb["Load Test"]

    _fill_header(ws, data)
    _fill_equipment_block(ws, data)

    devices = _devices_for_test(data, topo_devices)
    _ensure_block_count(ws, len(devices))
    for i, dev in enumerate(devices):
        _fill_block(ws, _FIRST_BLOCK_ROW + i * _BLOCK_STRIDE, dev, data, sim_summaries, use360)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_header(ws, data: dict) -> None:
    test = data.get("test", {}) or {}
    site = data.get("site", {}) or {}

    ws["J4"]  = test.get("name", "")
    ws["AM4"] = site.get("station", "")
    epoch = test.get("epoch")
    if epoch:
        ws["AX4"] = datetime.fromtimestamp(epoch, tz=timezone.utc).strftime("%Y-%m-%d")
    ws["AM5"] = test.get("created_by", "")

    ws["J17"]  = test.get("vref_label") or ""
    vref_mag = test.get("vref_magnitude")
    if vref_mag is not None:
        ws["AL17"] = vref_mag
        ws["AL17"].number_format = "0.00"
        ws["AU17"] = 0
        ws["AU17"].number_format = "0.00"
    else:
        ws["AL17"] = ""
        ws["AU17"] = ""


def _fill_equipment_block(ws, data: dict) -> None:
    devices = sorted({
        did
        for sess in data.get("sessions", [])
        for did in (sess.get("by_device") or {}).keys()
    })
    drawings = data.get("drawings", []) or []
    parts: list[str] = ["Equipment:"]
    if devices:
        parts.extend(f"  - {did}" for did in devices)
    else:
        parts.append("  (none)")
    parts.append("")
    parts.append("Drawings:")
    if drawings:
        for d in drawings:
            title = (d.get("title") or "").strip() or "(untitled)"
            url = (d.get("url") or "").strip()
            rev = (d.get("revision") or "").strip()
            line = f"  - {title}"
            if rev: line += f" [rev {rev}]"
            if url: line += f"  {url}"
            parts.append(line)
    else:
        parts.append("  (none)")
    initial = (data.get("test", {}) or {}).get("description", "")
    parts.append("")
    parts.append("Initial Conditions:")
    parts.append(initial.strip() if initial else "")
    ws["A8"] = "\n".join(parts)


def _devices_for_test(data: dict, topo_devices: dict) -> list[dict]:
    test = data.get("test", {}) or {}
    sessions = data.get("sessions", [])
    
    ordered: list[str] = []
    seen = set()
    
    # 1. Use capture_points if defined
    cp_raw = test.get("capture_points")
    cp = []
    if cp_raw:
        try:
            cp = json.loads(cp_raw) if isinstance(cp_raw, str) else cp_raw
        except Exception: pass
    
    for did in cp:
        if did not in seen:
            seen.add(did)
            ordered.append(did)
            
    # 2. Add devices that already have measurements
    for sess in sorted(sessions, key=lambda s: s.get("epoch", 0)):
        for did in (sess.get("by_device") or {}).keys():
            if did not in seen:
                seen.add(did)
                ordered.append(did)

    devs: list[dict] = []
    for did in ordered:
        topo = topo_devices.get(did, {}) or {}
        dtype = topo.get("type") or _infer_type_from_measurements(data, did)

        if dtype == "Relay":
            inputs = _relay_inputs(did, topo_devices)
            if inputs:
                for inp_id in inputs:
                    inp_topo = topo_devices.get(inp_id, {}) or {}
                    inp_type = inp_topo.get("type", "")
                    inp_ratio_text = (inp_topo.get("ratio") or inp_topo.get("selected_tap") or "").strip()
                    inp_ratio_factor = _ratio_to_factor(inp_ratio_text) if inp_ratio_text else None
                    inp_bushing = (inp_topo.get("bushing") or "").strip()
                    inp_parent = _root_device_for(inp_id, inp_topo, topo_devices)
                    inp_root = f"{inp_parent} {inp_bushing}".strip() if inp_parent or inp_bushing else ""
                    devs.append({
                        "id": did,
                        "type": "Relay",
                        "name": did,
                        "location": (inp_topo.get("location") or "").strip(),
                        "ratio": inp_ratio_text,
                        "ratio_factor": inp_ratio_factor,
                        "root_device": inp_root,
                        "kind": _kind_for_type(inp_type),
                        "is_relay": True,
                        "vt_ratio": inp_topo.get("ratio") if "Voltage" in inp_type else "",
                    })
                continue

        ratio_text = (topo.get("ratio") or topo.get("selected_tap") or "").strip()
        ratio_factor = _ratio_to_factor(ratio_text) if ratio_text else None
        bushing = (topo.get("bushing") or "").strip()
        parent = _root_device_for(did, topo, topo_devices)
        root_device = f"{parent} {bushing}".strip() if parent or bushing else ""
        devs.append({
            "id": did,
            "type": dtype,
            "name": did,
            "location": (topo.get("location") or "").strip(),
            "ratio": ratio_text,
            "ratio_factor": ratio_factor,
            "root_device": root_device,
            "kind": _kind_for_type(dtype),
            "vt_ratio": topo.get("ratio") if "Voltage" in (dtype or "") else "",
        })
    return devs


def _relay_inputs(relay_id: str, topo_devices: dict) -> list[str]:
    return [did for did, topo in topo_devices.items() if relay_id in (topo.get("secondary_connections") or [])]

def _ratio_to_factor(ratio_text: str) -> float | None:
    if not ratio_text: return None
    try:
        if ":" in ratio_text:
            a, b = ratio_text.split(":", 1)
            return float(a) / float(b)
        return float(ratio_text)
    except (ValueError, ZeroDivisionError): return None

def _root_device_for(did: str, topo: dict, topo_devices: dict) -> str:
    for key in ("parent", "host", "host_device", "downstream_device"):
        if topo.get(key): return str(topo[key])
    for other_id, other in topo_devices.items():
        for key in ("ct_ids", "vt_ids", "secondary_devices", "x_secondaries", "h_secondaries"):
            ids = other.get(key)
            if isinstance(ids, list) and did in ids: return other_id
    return ""

def _infer_type_from_measurements(data: dict, did: str) -> str:
    keys = set()
    for sess in data.get("sessions", []):
        keys.update(((sess.get("by_device") or {}).get(did) or {}).keys())
    if any("Voltage" in k or "V-Angle" in k for k in keys): return "VoltageTransformer"
    if any("Current" in k or "I-Angle" in k for k in keys): return "CurrentTransformer"
    return ""

def _kind_for_type(dtype: str) -> str:
    t = (dtype or "").lower()
    if any(x in t for x in ("voltagetransformer", "vt", "dualwindingvt", "ftblock", "isoblock", "voltage")):
        return "voltage"
    return "current"


def _fill_block(ws, r: int, dev: dict, data: dict, sim_summaries: dict, use360: bool) -> None:
    stype = _short_type(dev["type"])
    ws.cell(row=r, column=1).value = stype
    ws[f"G{r}"] = dev["name"]
    ws[f"G{r+1}"] = dev["root_device"]
    ws[f"C{r+2}"] = ""
    ws[f"L{r+2}"] = dev["ratio"]
    ws[f"J{r+3}"] = ""
    
    if dev.get("vt_ratio"):
        ws[f"L{r+2}"] = dev["vt_ratio"]

    # Coloring
    fill = None
    if stype == "Relay": fill = _BLUE_FILL
    elif stype == "CTTB": fill = _YELLOW_FILL
    elif stype in ("FT", "ISO"): fill = _RED_FILL
    elif dev["kind"] == "voltage": fill = _RED_FILL
    elif dev["kind"] == "current": fill = _BLUE_FILL

    if fill:
        for dr in range(_BLOCK_STRIDE):
            for dc in range(1, 57): ws.cell(row=r+dr, column=dc).fill = fill

    keys = _VOLTAGE_KEYS if dev["kind"] == "voltage" else _CURRENT_KEYS
    measurements = _latest_measurements_for_device(data, dev["id"])
    ratio = dev["ratio_factor"]
    sim_summary = copy(sim_summaries.get(dev["id"], {}))

    # S&I Logic: detect isolated phases from measurements
    measured_mags = {p: measurements.get(keys[p][0], {}).get("value") for p in ["A", "B", "C"]}
    has_any_meas = any(m is not None for m in measured_mags.values())
    if has_any_meas:
        threshold = 0.05
        active_phases = [p for p, m in measured_mags.items() if m is not None and m > threshold]
        if active_phases and len(active_phases) < 3:
            isolated = [p for p in ["A", "B", "C"] if p not in active_phases]
            ws[f"C{r+2}"] = "Isolated: " + ", ".join(isolated)
            # Adjust predictions to match S&I state
            for p in isolated:
                sim_summary[keys[p][0]] = 0
                sim_summary[keys[p][1]] = 0

    for phase, (label_col, mag_col, ang_col) in _PHASE_COLS.items():
        mag_key, ang_key = keys[phase]
        
        # Measured Secondary
        mv = measurements.get(mag_key, {})
        mag = mv.get("value")
        ang = _format_angle(measurements.get(ang_key, {}).get("value"), use360)

        # Predicted Secondary (from simulation summary, possibly zeroed above)
        pred_mag = sim_summary.get(mag_key)
        pred_ang = _format_angle(sim_summary.get(ang_key), use360)

        ws[f"{label_col}{r}"] = _phase_label(phase, dev["kind"])
        
        # Row r+1: Measured Secondary
        if mag is not None:
            c = ws[f"{mag_col}{r+1}"]
            c.value = mag; c.number_format = "0.00"
            if ang is not None:
                c = ws[f"{ang_col}{r+1}"]
                c.value = ang; c.number_format = "0.00"
        
        # Row r+2: Predicted Secondary
        if pred_mag is not None:
            c = ws[f"{mag_col}{r+2}"]
            c.value = pred_mag; c.number_format = "0.00"
            if pred_ang is not None:
                c = ws[f"{ang_col}{r+2}"]
                c.value = pred_ang; c.number_format = "0.00"

        # Row r+3: Measured Primary
        if mag is not None and ratio:
            c = ws[f"{mag_col}{r+3}"]
            c.value = mag * ratio; c.number_format = "0.00"
            if ang is not None:
                c = ws[f"{ang_col}{r+3}"]
                c.value = ang; c.number_format = "0.00"


def _short_type(dtype: str) -> str:
    t = (dtype or "").lower()
    if "currenttransformer" in t: return "CT"
    if "voltagetransformer" in t or "dualwindingvt" in t: return "VT"
    if "relay" in t: return "Relay"
    if "cttb" in t: return "CTTB"
    if "ftblock" in t: return "FT"
    if "isoblock" in t: return "ISO"
    return dtype or ""

def _phase_label(phase: str, kind: str) -> str:
    if kind == "voltage": return "" if phase == "N" else f"{phase} to N"
    return phase

def _latest_measurements_for_device(data: dict, device_id: str) -> dict:
    result = {}
    for sess in data.get("sessions", []):
        for key, mv in (sess.get("by_device") or {}).get(device_id, {}).items():
            if key not in result or (mv.get("epoch") or 0) > result[key].get("epoch", 0):
                result[key] = mv
    return result


# ── Ingestion ─────────────────────────────────────────────────────────────────

def ingest_load_test_report(db_path: str, test_id: str, xlsx_bytes: bytes) -> str | None:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Load Test"]
    tech = ws["AM5"].value or "Excel Import"
    sess_id = site_db.start_session(db_path, label="Excel Manual Entry", instrument="manual", technician=tech, test_id=test_id)
    r = _FIRST_BLOCK_ROW
    while r < ws.max_row:
        dev_name = ws[f"G{r}"].value
        if not dev_name: 
            r += _BLOCK_STRIDE
            continue
        dtype_short = ws[f"A{r}"].value
        if not dtype_short:
            r += _BLOCK_STRIDE
            continue
        kind = "voltage" if "VT" in str(dtype_short) else "current"
        keys = _VOLTAGE_KEYS if kind == "voltage" else _CURRENT_KEYS
        measurements = {}
        for phase, (_, mag_col, ang_col) in _PHASE_COLS.items():
            mag = ws[f"{mag_col}{r+1}"].value
            ang = ws[f"{ang_col}{r+1}"].value
            if isinstance(mag, (int, float)): measurements[keys[phase][0]] = mag
            if isinstance(ang, (int, float)): measurements[keys[phase][1]] = ang
        if measurements:
            site_db.record_measurements(db_path, sess_id, dev_name, measurements)
        r += _BLOCK_STRIDE
    return sess_id


# ── Dynamic Block Count ───────────────────────────────────────────────────────

def _ensure_block_count(ws, n: int) -> None:
    n = max(n, 1)
    if n == _TEMPLATE_BLOCK_COUNT: return
    if n < _TEMPLATE_BLOCK_COUNT:
        _delete_rows_with_merges(ws, _FIRST_BLOCK_ROW + n * _BLOCK_STRIDE, _BLOCK_STRIDE * (_TEMPLATE_BLOCK_COUNT - n))
        return
    extra = n - _TEMPLATE_BLOCK_COUNT
    src_first = _FIRST_BLOCK_ROW + (_TEMPLATE_BLOCK_COUNT - 1) * _BLOCK_STRIDE
    insert_at = src_first + _BLOCK_STRIDE
    src_cells = [[_snapshot_cell(ws.cell(row=src_first+dr, column=c)) for c in range(1, ws.max_column+1)] for dr in range(_BLOCK_STRIDE)]
    src_merges = [(rng.min_row-src_first, rng.max_row-src_first, rng.min_col, rng.max_col) for rng in ws.merged_cells.ranges if src_first <= rng.min_row < src_first+_BLOCK_STRIDE]
    _insert_rows_with_merges(ws, insert_at, _BLOCK_STRIDE * extra)
    for blk in range(extra):
        dst = insert_at + blk * _BLOCK_STRIDE
        for dr in range(_BLOCK_STRIDE):
            for ci, snap in enumerate(src_cells[dr], 1): _restore_cell(ws.cell(row=dst+dr, column=ci), snap)
        for r0, r1, c0, c1 in src_merges: ws.merge_cells(start_row=dst+r0, end_row=dst+r1, start_column=c0, end_column=c1)

def _insert_rows_with_merges(ws, idx, amount):
    affected = [rng for rng in list(ws.merged_cells.ranges) if rng.min_row >= idx]
    for rng in affected: ws.unmerge_cells(str(rng))
    ws.insert_rows(idx, amount)
    for rng in affected: ws.merge_cells(start_row=rng.min_row+amount, end_row=rng.max_row+amount, start_column=rng.min_col, end_column=rng.max_col)

def _delete_rows_with_merges(ws, idx, amount):
    end = idx + amount
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
        if rng.max_row < idx: nmin, nmax = rng.min_row, rng.max_row
        elif rng.min_row >= end: nmin, nmax = rng.min_row-amount, rng.max_row-amount
        elif rng.min_row < idx <= rng.max_row < end: nmin, nmax = rng.min_row, idx-1
        elif idx <= rng.min_row < rng.max_row < end: continue
        else: nmin, nmax = min(rng.min_row, idx), rng.max_row-amount
        if nmax >= nmin: ws.merge_cells(start_row=nmin, end_row=nmax, start_column=rng.min_col, end_column=rng.max_col)
    ws.delete_rows(idx, amount)

def _snapshot_cell(c): return {"value": c.value, "font": copy(c.font), "border": copy(c.border), "fill": copy(c.fill), "alignment": copy(c.alignment), "number_format": c.number_format, "protection": copy(c.protection)}
def _restore_cell(c, s):
    c.value, c.font, c.border, c.fill, c.alignment, c.number_format, c.protection = s["value"], s["font"], s["border"], s["fill"], s["alignment"], s["number_format"], s["protection"]

def _format_angle(deg: float, use360: bool) -> float | None:
    if deg is None: return None
    if use360:
        return (deg % 360 + 360) % 360
    else:
        return (deg + 180) % 360 - 180
