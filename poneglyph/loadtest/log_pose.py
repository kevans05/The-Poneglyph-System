"""Log Pose — P&C Equipment Load Test wizard panel."""
from __future__ import annotations

import csv
import uuid
from dataclasses import asdict
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Optional
import tkinter as tk
from tkinter import ttk

from .sea_chart import MeasurementPoint, VectorGroup, LoadTestRecord
from .vivre_card import check_forcing_neutral, differential_deviation


# ── Connectivity helpers ──────────────────────────────────────────────────────

def _build_adj(diagram) -> dict[str, set[str]]:
    """Undirected device adjacency map built from relay wires."""
    adj: dict[str, set[str]] = {}
    for rw in diagram.get_relay_wires().values():
        s, d = rw.source_id, rw.relay_id
        adj.setdefault(s, set()).add(d)
        adj.setdefault(d, set()).add(s)
    return adj


def _full_path(adj: dict[str, set[str]], start: str) -> set[str]:
    """BFS all devices reachable from start (start excluded from result)."""
    visited = {start}
    queue = [start]
    while queue:
        node = queue.pop(0)
        for nb in adj.get(node, set()):
            if nb not in visited:
                visited.add(nb)
                queue.append(nb)
    visited.discard(start)
    return visited


# ── Main panel ────────────────────────────────────────────────────────────────

class LogPosePanel(tk.Frame):
    """P&C Load Test panel: wizard-driven test workflow + History."""

    _STEP_TITLES = [
        "Job Info",             # 0
        "New or Repeat",        # 1
        "Device Selection",     # 2
        "Drawings",             # 3
        "Entry Method",         # 4
        "Protection Blocking",  # 5  (Manual path)
        "Download Template",    # 6  (Excel path)
        "Measurement Order",    # 7  (Manual path)
        "Upload Data",          # 8  (Excel path)
        "Measurements",         # 9
    ]
    _CRUMB_ABBREV = [
        "Job", "New/Rep", "Devices", "Drawings",
        "Method", "Blocking", "Download", "Order", "Upload", "Measure",
    ]

    def __init__(self, parent: tk.Widget, diagram=None, project_path=None,
                 project_name: str = "") -> None:
        super().__init__(parent)
        self._diagram = diagram
        self._project_path = project_path
        self._project_name = project_name

        self._points: list[MeasurementPoint] = []
        self._blocked = tk.BooleanVar(value=False)

        # Job info vars
        self._tech_var   = tk.StringVar()
        self._email_var  = tk.StringVar()
        self._wo_var     = tk.StringVar()
        self._note_var   = tk.StringVar()

        # Wizard routing state
        self._wiz_is_new        = True
        self._wiz_manual        = True
        self._wiz_repeat_record: Optional[dict] = None
        self._wiz_selected_ids: list[str] = []
        self._wiz_test_drawings: list[str] = []
        self._wiz_suggested_ids: list[str] = []
        self._wiz_upload_path   = ""

        # Measurement grid vars
        self._meas_vars: dict[int, dict[str, tk.StringVar]] = {}

        # Vector analysis vars
        self._vec_a_mag = tk.StringVar(value="0.0")
        self._vec_a_ang = tk.StringVar(value="0.0")
        self._vec_b_mag = tk.StringVar(value="0.0")
        self._vec_b_ang = tk.StringVar(value="120.0")
        self._vec_c_mag = tk.StringVar(value="0.0")
        self._vec_c_ang = tk.StringVar(value="240.0")
        self._vec_n_mag = tk.StringVar(value="0.0")
        self._vec_n_ang = tk.StringVar(value="0.0")
        self._vec_result_var = tk.StringVar(value="—")

        # Forcing neutral vars
        self._fn_a_mag = tk.StringVar(value="0.0")
        self._fn_a_ang = tk.StringVar(value="0.0")
        self._fn_n_mag = tk.StringVar(value="0.0")
        self._fn_n_ang = tk.StringVar(value="180.0")
        self._fn_result_var = tk.StringVar(value="—")

        # Differential wizard vars
        self._diff_cttbs: list[str] = []
        self._diff_step  = 0
        self._diff_points: list[MeasurementPoint] = []
        self._diff_mag_var = tk.StringVar(value="0.0")
        self._diff_ang_var = tk.StringVar(value="0.0")

        self._build_ui()

        if self._diagram is not None:
            self._load_from_sld()

    # ── Top-level UI ──────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        # Ensure treeview rows are tall enough that text isn't clipped on any platform
        style = ttk.Style()
        style.configure("Treeview", rowheight=26)
        style.configure("Treeview.Heading", font=("TkDefaultFont", 9, "bold"))

        self._outer_nb = ttk.Notebook(self)
        self._outer_nb.pack(fill=tk.BOTH, expand=True)

        hist_frame = tk.Frame(self._outer_nb)
        self._outer_nb.add(hist_frame, text="History")
        self._build_history_tab(hist_frame)

        test_frame = tk.Frame(self._outer_nb)
        self._outer_nb.add(test_frame, text="Active Test")
        self._build_wizard(test_frame)

        self._outer_nb.bind("<<NotebookTabChanged>>", self._on_outer_tab_changed)

    # ── Wizard shell ──────────────────────────────────────────────────────────

    def _build_wizard(self, parent: tk.Frame) -> None:
        self._wiz_current = 0

        # Header bar
        hdr = tk.Frame(parent, bg="#1A252F")
        hdr.pack(fill=tk.X)
        self._wiz_title_lbl = tk.Label(
            hdr, text="", font=("TkDefaultFont", 11, "bold"),
            fg="white", bg="#1A252F", anchor="w",
        )
        self._wiz_title_lbl.pack(side=tk.LEFT, padx=14, pady=8)
        self._wiz_crumb_frame = tk.Frame(hdr, bg="#1A252F")
        self._wiz_crumb_frame.pack(side=tk.RIGHT, padx=14)

        # Content area (step frames stacked here)
        self._wiz_content = tk.Frame(parent)
        self._wiz_content.pack(fill=tk.BOTH, expand=True)

        # Nav bar
        nav = tk.Frame(parent, bd=1, relief="ridge")
        nav.pack(fill=tk.X, side=tk.BOTTOM)
        self._wiz_back_btn = tk.Button(nav, text="◀  Back", command=self._wiz_back, width=10)
        self._wiz_back_btn.pack(side=tk.LEFT, padx=14, pady=5)
        self._wiz_next_btn = tk.Button(nav, text="Next  ▶", command=self._wiz_next, width=10)
        self._wiz_next_btn.pack(side=tk.RIGHT, padx=14, pady=5)

        # Build all step frames
        self._wiz_frames: dict[int, tk.Frame] = {}
        builders = [
            self._build_step_job_info,
            self._build_step_new_or_repeat,
            self._build_step_devices,
            self._build_step_drawings,
            self._build_step_entry_method,
            self._build_step_blocking,
            self._build_step_download,
            self._build_step_order,
            self._build_step_upload,
            self._build_step_measurements,
        ]
        for i, builder in enumerate(builders):
            f = tk.Frame(self._wiz_content)
            builder(f)
            self._wiz_frames[i] = f

        self._wiz_show(0)

    # ── Wizard navigation ─────────────────────────────────────────────────────

    def _active_route(self) -> list[int]:
        route = [0, 1]
        if self._wiz_is_new:
            route.append(2)
        route.append(3)
        route.append(4)
        if self._wiz_manual:
            route.extend([5, 7])
        else:
            route.extend([6, 8])
        route.append(9)
        return route

    def _wiz_show(self, step: int) -> None:
        for f in self._wiz_frames.values():
            f.pack_forget()
        self._wiz_frames[step].pack(fill=tk.BOTH, expand=True)
        self._wiz_current = step
        self._wiz_update_header()
        self._wiz_update_nav()
        self._wiz_on_enter(step)

    def _wiz_on_enter(self, step: int) -> None:
        """Side effects when a wizard step becomes active."""
        if step == 2:
            self._populate_dev_tree()
            self._refresh_sel_listbox()
        elif step == 3:
            self._refresh_drw_tree()
        elif step == 6:
            self._sync_points_from_selection()
        elif step == 7:
            self._sync_points_from_selection()
            self._refresh_order_list()
        elif step == 9:
            self._sync_points_from_selection()
            self._refresh_measurements_tab()

    def _wiz_update_header(self) -> None:
        step  = self._wiz_current
        route = self._active_route()
        pos   = (route.index(step) + 1) if step in route else 1
        total = len(route)
        self._wiz_title_lbl.config(
            text=f"Step {pos} of {total}:  {self._STEP_TITLES[step]}"
        )
        for w in self._wiz_crumb_frame.winfo_children():
            w.destroy()
        for i, s in enumerate(route):
            is_current = (s == step)
            is_done    = (route.index(s) < route.index(step)) if step in route else False
            fg   = "white" if is_current else ("#27AE60" if is_done else "#7F8C8D")
            font = ("TkDefaultFont", 8, "bold") if is_current else ("TkDefaultFont", 8)
            abbrev = self._CRUMB_ABBREV[s] if s < len(self._CRUMB_ABBREV) else self._STEP_TITLES[s]
            tk.Label(self._wiz_crumb_frame, text=abbrev,
                     fg=fg, bg="#1A252F", font=font).pack(side=tk.LEFT)
            if i < len(route) - 1:
                tk.Label(self._wiz_crumb_frame, text=" › ",
                         fg="#555", bg="#1A252F",
                         font=("TkDefaultFont", 8)).pack(side=tk.LEFT)

    def _wiz_update_nav(self) -> None:
        step  = self._wiz_current
        route = self._active_route()
        pos   = route.index(step) if step in route else 0
        self._wiz_back_btn.config(state="normal" if pos > 0 else "disabled")
        at_end = (pos == len(route) - 1)
        self._wiz_next_btn.config(text="Finish" if at_end else "Next  ▶")

    def _wiz_validate(self, step: int) -> bool:
        if step == 0:
            return bool(self._tech_var.get().strip() and self._wo_var.get().strip())
        if step == 1 and not self._wiz_is_new:
            return self._wiz_repeat_record is not None
        if step == 5:
            return self._blocked.get()
        return True

    def _wiz_next(self) -> None:
        step = self._wiz_current
        if not self._wiz_validate(step):
            self._wiz_validation_msg(step)
            return
        route = self._active_route()
        pos   = route.index(step) if step in route else 0
        if pos < len(route) - 1:
            self._wiz_show(route[pos + 1])
        else:
            self._save_test()

    def _wiz_back(self) -> None:
        route = self._active_route()
        step  = self._wiz_current
        pos   = route.index(step) if step in route else 0
        if pos > 0:
            self._wiz_show(route[pos - 1])

    def _wiz_validation_msg(self, step: int) -> None:
        if step == 0:
            messagebox.showwarning("Required",
                "Please enter your name and WO number before continuing.", parent=self)
        elif step == 1:
            messagebox.showwarning("Required",
                "Select a previous test from the list to repeat.", parent=self)
        elif step == 5:
            messagebox.showwarning("Blocking required",
                "You must confirm protection blocking before proceeding.", parent=self)

    # ── Step 0: Job Info ──────────────────────────────────────────────────────

    def _build_step_job_info(self, parent: tk.Frame) -> None:
        canvas = tk.Canvas(parent)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inner = tk.Frame(canvas)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win_id, width=canvas.winfo_width())

        inner.bind("<Configure>", _cfg)
        canvas.bind("<Configure>", _cfg)

        pad = {"padx": 16, "pady": 6}

        tk.Label(inner, text="Who is performing this test?",
                 font=("TkDefaultFont", 12, "bold")).pack(anchor="w", **pad)

        frm = tk.LabelFrame(inner, text="Technologist", padx=10, pady=8)
        frm.pack(fill=tk.X, **pad)

        for row, (lbl, var) in enumerate([
            ("Full Name *",           self._tech_var),
            ("Email",                 self._email_var),
            ("WO / Project Number *", self._wo_var),
        ]):
            tk.Label(frm, text=lbl, anchor="e", width=24).grid(
                row=row, column=0, sticky="e", pady=4)
            tk.Entry(frm, textvariable=var, width=36).grid(
                row=row, column=1, sticky="w", padx=8)

        tk.Label(inner, text="* required to continue",
                 fg="#888", font=("TkDefaultFont", 8)).pack(anchor="w", padx=16)

    # ── Step 1: New or Repeat ─────────────────────────────────────────────────

    def _build_step_new_or_repeat(self, parent: tk.Frame) -> None:
        self._wiz_mode_var = tk.StringVar(value="new")

        top = tk.Frame(parent)
        top.pack(fill=tk.BOTH, expand=True, padx=20, pady=16)

        tk.Label(top, text="What would you like to do?",
                 font=("TkDefaultFont", 12, "bold")).pack(anchor="w", pady=(0, 14))

        def _option(value, title, desc):
            frm = tk.Frame(top, bd=1, relief="groove", padx=12, pady=8)
            frm.pack(fill=tk.X, pady=3)
            tk.Radiobutton(
                frm, text=title, variable=self._wiz_mode_var, value=value,
                font=("TkDefaultFont", 10, "bold"),
                command=self._wiz_on_mode_change,
            ).pack(anchor="w")
            tk.Label(frm, text=desc, fg="#555", justify="left",
                     wraplength=480, font=("TkDefaultFont", 9)).pack(
                anchor="w", padx=(22, 0), pady=(2, 0))

        _option("new",    "New Test",
                "Start fresh — pick devices, drawings, and enter measurements.")
        _option("repeat", "Repeat Previous",
                "Re-run an existing test. Device list and order are carried over; measurements are cleared.")

        # History picker — shown only when "repeat" is selected
        self._repeat_picker = tk.Frame(top)

        tk.Label(self._repeat_picker, text="Select test to repeat:",
                 font=("TkDefaultFont", 9, "bold")).pack(anchor="w", pady=(10, 2))

        search_row = tk.Frame(self._repeat_picker)
        search_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(search_row, text="Search:").pack(side=tk.LEFT)
        self._repeat_search_var = tk.StringVar()
        self._repeat_search_var.trace_add("write", lambda *_: self._repeat_filter())
        tk.Entry(search_row, textvariable=self._repeat_search_var, width=30).pack(
            side=tk.LEFT, padx=6)

        cols = ("timestamp", "wo_number", "technologist")
        self._repeat_tree = ttk.Treeview(
            self._repeat_picker, columns=cols, show="headings",
            selectmode="browse", height=8,
        )
        self._repeat_tree.heading("timestamp",    text="Date / Time")
        self._repeat_tree.heading("wo_number",    text="WO #")
        self._repeat_tree.heading("technologist", text="Technologist")
        self._repeat_tree.column("timestamp",    width=160)
        self._repeat_tree.column("wo_number",    width=120)
        self._repeat_tree.column("technologist", width=140)
        vsb = ttk.Scrollbar(self._repeat_picker, orient="vertical",
                             command=self._repeat_tree.yview)
        self._repeat_tree.configure(yscrollcommand=vsb.set)
        self._repeat_tree.pack(side=tk.LEFT, fill=tk.X, expand=True)
        vsb.pack(side=tk.LEFT, fill=tk.Y)
        self._repeat_tree.bind("<<TreeviewSelect>>", self._on_repeat_select)
        self._all_history_records: list[dict] = []

    def _wiz_on_mode_change(self) -> None:
        is_repeat = (self._wiz_mode_var.get() == "repeat")
        self._wiz_is_new = not is_repeat
        if is_repeat:
            self._repeat_picker.pack(fill=tk.BOTH, expand=True, pady=(8, 0))
            self._repeat_load_history()
        else:
            self._repeat_picker.pack_forget()
            self._wiz_repeat_record = None
        self._wiz_update_header()
        self._wiz_update_nav()

    def _repeat_load_history(self) -> None:
        if not self._project_path:
            return
        try:
            from poneglyph.io.project import list_load_tests
            self._all_history_records = list_load_tests(Path(self._project_path))
        except Exception:
            self._all_history_records = []
        self._repeat_filter()

    def _repeat_filter(self) -> None:
        query = self._repeat_search_var.get().lower()
        tree  = self._repeat_tree
        for item in tree.get_children():
            tree.delete(item)
        for rec in self._all_history_records:
            ts   = rec.get("timestamp", "")[:19].replace("T", " ")
            wo   = rec.get("wo_number", "")
            tech = rec.get("technologist", "")
            if query and query not in (ts + wo + tech).lower():
                continue
            tree.insert("", tk.END, iid=rec.get("id", ""), values=(ts, wo, tech))

    def _on_repeat_select(self, _event) -> None:
        sel = self._repeat_tree.selection()
        if not sel:
            return
        rec = next((r for r in self._all_history_records
                    if r.get("id") == sel[0]), None)
        if rec is None:
            return
        self._wiz_repeat_record = rec
        # Pre-fill job info
        self._tech_var.set(rec.get("technologist", ""))
        self._wo_var.set(rec.get("wo_number", ""))
        self._note_var.set(rec.get("blocking_note", ""))
        # Restore device list — clear measurements
        self._points.clear()
        valid_fields = set(MeasurementPoint.__dataclass_fields__)
        for p in rec.get("points", []):
            self._points.append(MeasurementPoint(**{k: v for k, v in p.items()
                                                    if k in valid_fields}))
        for pt in self._points:
            pt.meas_magnitude = 0.0
            pt.meas_angle     = 0.0
        self._wiz_selected_ids  = [p.device_id for p in self._points]
        self._wiz_test_drawings = list(rec.get("drawings", []))

    # ── Step 2: Device Selection ──────────────────────────────────────────────

    def _build_step_devices(self, parent: tk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(2, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Available devices:",
                 font=("TkDefaultFont", 9, "bold")).grid(
            row=0, column=0, sticky="w", padx=(12, 0), pady=(8, 2))
        tk.Label(parent, text="Test devices:",
                 font=("TkDefaultFont", 9, "bold")).grid(
            row=0, column=2, sticky="w", padx=(4, 12), pady=(8, 2))

        # Left: tree of all SLD devices
        left = tk.Frame(parent, bd=1, relief="sunken")
        left.grid(row=1, column=0, sticky="nsew", padx=(12, 0), pady=(0, 8))
        left.rowconfigure(0, weight=1)
        left.columnconfigure(0, weight=1)
        self._dev_tree = ttk.Treeview(left, show="tree", selectmode="extended")
        vsb_l = ttk.Scrollbar(left, orient="vertical", command=self._dev_tree.yview)
        self._dev_tree.configure(yscrollcommand=vsb_l.set)
        self._dev_tree.grid(row=0, column=0, sticky="nsew", pady=(2, 0))
        vsb_l.grid(row=0, column=1, sticky="ns")
        self._dev_tree.bind("<<TreeviewSelect>>", self._on_dev_tree_select)

        # Centre: transfer buttons
        mid = tk.Frame(parent)
        mid.grid(row=1, column=1, padx=6)
        tk.Button(mid, text="Add →",    command=self._dev_add_selected,  width=8).pack(pady=6)
        tk.Button(mid, text="← Remove", command=self._dev_remove_selected, width=8).pack(pady=6)

        # Right: selected devices listbox
        right = tk.Frame(parent, bd=1, relief="sunken")
        right.grid(row=1, column=2, sticky="nsew", padx=(0, 12), pady=(0, 8))
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)
        self._sel_listbox = tk.Listbox(right, selectmode=tk.EXTENDED)
        vsb_r = ttk.Scrollbar(right, orient="vertical", command=self._sel_listbox.yview)
        self._sel_listbox.configure(yscrollcommand=vsb_r.set)
        self._sel_listbox.grid(row=0, column=0, sticky="nsew")
        vsb_r.grid(row=0, column=1, sticky="ns")

        # Suggestion banner (hidden until a connected device is found)
        self._sug_frame = tk.Frame(parent, bg="#FFF8E1", bd=1, relief="ridge")
        self._sug_lbl = tk.Label(self._sug_frame, text="", bg="#FFF8E1",
                                  wraplength=400, justify="left", anchor="w")
        self._sug_lbl.pack(side=tk.LEFT, padx=8, pady=6, fill=tk.X, expand=True)
        tk.Button(self._sug_frame, text="Add All Connected",
                  command=self._dev_add_suggested).pack(side=tk.RIGHT, padx=6)
        tk.Button(self._sug_frame, text="✕",
                  command=lambda: self._sug_frame.grid_remove()).pack(side=tk.RIGHT, padx=(6, 0))

        # Suggestion banner goes in row 2 spanning all columns
        # (grid_remove hides it; grid() shows it)

    def _populate_dev_tree(self) -> None:
        tree = self._dev_tree
        for item in tree.get_children():
            tree.delete(item)
        if self._diagram is None:
            return
        # CTs and VTs are source devices — not test targets
        for group_name, attr in [
            ("CTTBs",       "_cttbs"),
            ("Test Blocks", "_testblocks"),
            ("Relays",      "_relays"),
        ]:
            dev_dict = getattr(self._diagram, attr, {})
            if not dev_dict:
                continue
            pid = tree.insert("", tk.END, text=group_name, open=True, tags=("group",))
            for dev_id, dev in dev_dict.items():
                tree.insert(pid, tk.END, text=getattr(dev, "name", dev_id),
                            iid=dev_id, tags=("device",))
        tree.yview_moveto(0)

    def _on_dev_tree_select(self, _event) -> None:
        sel = self._dev_tree.selection()
        device_ids = [s for s in sel
                      if "group" not in self._dev_tree.item(s, "tags")]
        if not device_ids or self._diagram is None:
            self._sug_frame.grid_remove()
            return

        adj       = _build_adj(self._diagram)
        connected: set[str] = set()
        for did in device_ids:
            connected |= _full_path(adj, did)

        # Exclude CTs/VTs (source devices, not test targets) and already-selected ids
        source_ids = (set(getattr(self._diagram, "_cts",  {}).keys()) |
                      set(getattr(self._diagram, "_vts",  {}).keys()))
        excluded   = set(self._wiz_selected_ids) | set(device_ids) | source_ids
        suggestions = connected - excluded
        if not suggestions:
            self._sug_frame.grid_remove()
            return

        all_devs = self._all_devices()
        names    = [getattr(all_devs.get(s), "name", s) for s in suggestions]
        self._wiz_suggested_ids = list(suggestions)
        self._sug_lbl.config(
            text=f"Connected devices found: {', '.join(names)}"
        )
        self._sug_frame.grid(row=2, column=0, columnspan=3, sticky="ew",
                              padx=12, pady=(0, 8))

    def _dev_add_selected(self) -> None:
        for dev_id in self._dev_tree.selection():
            if "group" not in self._dev_tree.item(dev_id, "tags"):
                if dev_id not in self._wiz_selected_ids:
                    self._wiz_selected_ids.append(dev_id)
        self._refresh_sel_listbox()
        self._sug_frame.grid_remove()

    def _dev_add_suggested(self) -> None:
        for dev_id in self._wiz_suggested_ids:
            if dev_id not in self._wiz_selected_ids:
                self._wiz_selected_ids.append(dev_id)
        for dev_id in self._dev_tree.selection():
            if "group" not in self._dev_tree.item(dev_id, "tags"):
                if dev_id not in self._wiz_selected_ids:
                    self._wiz_selected_ids.append(dev_id)
        self._refresh_sel_listbox()
        self._sug_frame.grid_remove()

    def _dev_remove_selected(self) -> None:
        for i in sorted(self._sel_listbox.curselection(), reverse=True):
            if 0 <= i < len(self._wiz_selected_ids):
                self._wiz_selected_ids.pop(i)
        self._refresh_sel_listbox()

    def _refresh_sel_listbox(self) -> None:
        all_devs = self._all_devices()
        self._sel_listbox.delete(0, tk.END)
        for dev_id in self._wiz_selected_ids:
            dev  = all_devs.get(dev_id)
            name = getattr(dev, "name", dev_id) if dev else dev_id
            self._sel_listbox.insert(tk.END, name)

    def _all_devices(self) -> dict:
        if self._diagram is None:
            return {}
        d: dict = {}
        for attr in ("_cts", "_vts", "_cttbs", "_testblocks", "_relays"):
            d.update(getattr(self._diagram, attr, {}))
        return d

    def _device_type(self, dev_id: str) -> str:
        for attr, dtype in [("_cts","ct"), ("_vts","vt"), ("_cttbs","cttb"),
                             ("_testblocks","testblock"), ("_relays","relay")]:
            if dev_id in getattr(self._diagram, attr, {}):
                return dtype
        return "unknown"

    def _sync_points_from_selection(self) -> None:
        """Build self._points from _wiz_selected_ids, preserving existing measurements."""
        if not self._wiz_is_new:
            return  # Repeat path: points already loaded from record
        existing  = {pt.device_id: pt for pt in self._points}
        all_devs  = self._all_devices()
        new_points: list[MeasurementPoint] = []
        for dev_id in self._wiz_selected_ids:
            if dev_id in existing:
                new_points.append(existing[dev_id])
            else:
                dev = all_devs.get(dev_id)
                new_points.append(MeasurementPoint(
                    device_id=dev_id,
                    device_name=getattr(dev, "name", dev_id) if dev else dev_id,
                    device_type=self._device_type(dev_id),
                ))
        self._points = new_points

    # ── Step 3: Drawings ──────────────────────────────────────────────────────

    def _build_step_drawings(self, parent: tk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Drawings for this test:",
                 font=("TkDefaultFont", 10, "bold")).grid(
            row=0, column=0, sticky="w", padx=12, pady=(8, 2))

        cols = ("name", "device", "rev", "title")
        self._drw_tree = ttk.Treeview(parent, columns=cols, show="headings",
                                       selectmode="browse")
        self._drw_tree.heading("name",   text="Drawing Number")
        self._drw_tree.heading("device", text="Device")
        self._drw_tree.heading("rev",    text="Rev")
        self._drw_tree.heading("title",  text="Title / Description")
        self._drw_tree.column("name",   width=160)
        self._drw_tree.column("device", width=120)
        self._drw_tree.column("rev",    width=40)
        self._drw_tree.column("title",  width=280)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=self._drw_tree.yview)
        self._drw_tree.configure(yscrollcommand=vsb.set)
        self._drw_tree.grid(row=1, column=0, sticky="nsew", padx=(12, 0), pady=(0, 4))
        vsb.grid(row=1, column=1, sticky="ns", pady=(0, 4))

        btn_row = tk.Frame(parent)
        btn_row.grid(row=2, column=0, columnspan=2, sticky="ew", padx=12, pady=4)
        tk.Button(btn_row, text="Add Drawing…",    command=self._drw_add).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_row, text="Remove from Test", command=self._drw_remove).pack(side=tk.LEFT, padx=4)
        tk.Label(btn_row,
                 text="New drawings are added to the project drawing registry.",
                 fg="#888", font=("TkDefaultFont", 8)).pack(side=tk.LEFT, padx=10)

    def _refresh_drw_tree(self) -> None:
        tree = self._drw_tree
        for item in tree.get_children():
            tree.delete(item)
        if self._diagram is None:
            return

        all_devs = self._all_devices()
        registry = self._diagram.get_drawings()
        shown: set[str] = set()

        for dev_id in self._wiz_selected_ids:
            dev = all_devs.get(dev_id)
            if dev is None:
                continue
            dev_name = getattr(dev, "name", dev_id)
            for drw_name in (getattr(dev, "device_drawings", None) or []):
                iid = f"{dev_id}::{drw_name}"
                if iid in shown:
                    continue
                shown.add(iid)
                drw   = registry.get(drw_name)
                rev   = getattr(drw, "rev", "") if drw else ""
                title = (getattr(drw, "title", "") or getattr(drw, "description", "")) if drw else ""
                tree.insert("", tk.END, iid=iid, values=(drw_name, dev_name, rev, title))
                if drw_name not in self._wiz_test_drawings:
                    self._wiz_test_drawings.append(drw_name)

        # Manually added drawings (not on any selected device)
        device_attached: set[str] = set()
        for dev_id in self._wiz_selected_ids:
            dev = all_devs.get(dev_id)
            for n in (getattr(dev, "device_drawings", None) or []):
                device_attached.add(n)

        for drw_name in self._wiz_test_drawings:
            if drw_name in device_attached:
                continue
            iid = f"manual::{drw_name}"
            if iid in shown:
                continue
            drw   = registry.get(drw_name)
            rev   = getattr(drw, "rev", "") if drw else ""
            title = (getattr(drw, "title", "") or getattr(drw, "description", "")) if drw else ""
            tree.insert("", tk.END, iid=iid, values=(drw_name, "—", rev, title))

    def _drw_add(self) -> None:
        if self._diagram is None:
            messagebox.showwarning("No diagram", "Open a diagram first.", parent=self)
            return
        dlg = tk.Toplevel(self)
        dlg.title("Add Drawing")
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)

        pad = {"padx": 10, "pady": 4}
        fields: dict[str, tk.StringVar] = {}
        for row, (lbl, key) in enumerate([
            ("Drawing Number:", "name"),
            ("Revision:",       "rev"),
            ("Title:",          "title"),
            ("URL:",            "url"),
        ]):
            tk.Label(dlg, text=lbl, anchor="e", width=16).grid(
                row=row, column=0, sticky="e", **pad)
            var = tk.StringVar()
            tk.Entry(dlg, textvariable=var, width=32).grid(
                row=row, column=1, sticky="w", **pad)
            fields[key] = var

        def _ok() -> None:
            name = fields["name"].get().strip()
            if not name:
                messagebox.showwarning("Required", "Drawing number is required.", parent=dlg)
                return
            from poneglyph.ui.diagram import DiagramDrawing
            if name not in self._diagram._drawings:
                self._diagram._drawings[name] = DiagramDrawing(
                    name=name,
                    rev=fields["rev"].get().strip(),
                    title=fields["title"].get().strip(),
                    description=fields["title"].get().strip(),
                    url=fields["url"].get().strip(),
                )
            if name not in self._wiz_test_drawings:
                self._wiz_test_drawings.append(name)
            dlg.destroy()
            self._refresh_drw_tree()

        btn_row = tk.Frame(dlg)
        btn_row.grid(row=4, column=0, columnspan=2, pady=8)
        tk.Button(btn_row, text="Add",    command=_ok).pack(side=tk.LEFT, padx=6)
        tk.Button(btn_row, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        dlg.wait_window()

    def _drw_remove(self) -> None:
        sel = self._drw_tree.selection()
        if not sel:
            return
        drw_name = sel[0].split("::")[-1]
        if drw_name in self._wiz_test_drawings:
            self._wiz_test_drawings.remove(drw_name)
        self._refresh_drw_tree()

    # ── Step 4: Entry Method ──────────────────────────────────────────────────

    def _build_step_entry_method(self, parent: tk.Frame) -> None:
        self._wiz_entry_var = tk.StringVar(value="manual")
        top = tk.Frame(parent)
        top.pack(fill=tk.BOTH, expand=True, padx=20, pady=16)

        tk.Label(top, text="How will you enter measurements?",
                 font=("TkDefaultFont", 12, "bold")).pack(anchor="w", pady=(0, 14))

        def _option(value, title, desc):
            frm = tk.Frame(top, bd=1, relief="groove", padx=12, pady=8)
            frm.pack(fill=tk.X, pady=3)
            tk.Radiobutton(
                frm, text=title, variable=self._wiz_entry_var, value=value,
                font=("TkDefaultFont", 10, "bold"),
                command=lambda v=value: self._wiz_set_manual(v == "manual"),
            ).pack(anchor="w")
            tk.Label(frm, text=desc, fg="#555", justify="left",
                     wraplength=480, font=("TkDefaultFont", 9)).pack(
                anchor="w", padx=(22, 0), pady=(2, 0))

        _option("manual", "Manual Entry",
                "Enter measurements step by step in the application.\n"
                "Requires protection blocking confirmation before measurements are unlocked.")
        _option("excel",  "Excel Spreadsheet",
                "Download a pre-filled template with device names and predicted values.\n"
                "Fill it in the field, then upload when back at the desk.")

    def _wiz_set_manual(self, is_manual: bool) -> None:
        self._wiz_manual = is_manual
        self._wiz_update_header()
        self._wiz_update_nav()

    # ── Step 5 (Manual): Protection Blocking ─────────────────────────────────

    def _build_step_blocking(self, parent: tk.Frame) -> None:
        canvas = tk.Canvas(parent)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inner = tk.Frame(canvas)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win_id, width=canvas.winfo_width())

        inner.bind("<Configure>", _cfg)
        canvas.bind("<Configure>", _cfg)
        pad = {"padx": 16, "pady": 6}

        block_frame = tk.LabelFrame(
            inner, text="⚠  Protection Blocking",
            fg="red", font=("TkDefaultFont", 10, "bold"),
            padx=10, pady=8, relief="ridge", bd=3,
        )
        block_frame.pack(fill=tk.X, **pad)

        tk.Label(block_frame, text="⚠  Protection Blocking Required",
                 font=("TkDefaultFont", 11, "bold"), fg="red").pack(anchor="w", pady=(0, 4))
        tk.Label(
            block_frame,
            text="Confirm FVO/SIO contact and SEL Relay Mirrored Bits blocked before proceeding.",
            wraplength=520, justify="left",
        ).pack(anchor="w")
        tk.Label(block_frame,
                 text="Blocking Note (who contacted, reference #):").pack(anchor="w", pady=(8, 2))
        tk.Entry(block_frame, textvariable=self._note_var, width=60).pack(anchor="w", fill=tk.X)
        tk.Checkbutton(
            block_frame,
            text="Protection has been blocked — I confirm",
            variable=self._blocked,
            font=("TkDefaultFont", 10, "bold"), fg="darkred",
            command=self._on_blocking_changed,
        ).pack(anchor="w", pady=(10, 0))

        meter_frame = tk.LabelFrame(inner, text="Meter Configuration Reminders",
                                     bg="#FFFDE7", padx=8, pady=6)
        meter_frame.pack(fill=tk.X, **pad)
        for r in [
            "• Connect reference lead to Channel 1",
            "• Set Channel 2 to lag by 0°–360°",
            "• Red side of current jack towards CT",
        ]:
            tk.Label(meter_frame, text=r, bg="#FFFDE7", anchor="w").pack(anchor="w", pady=1)

    # ── Step 6 (Excel): Download Template ────────────────────────────────────

    def _build_step_download(self, parent: tk.Frame) -> None:
        f = tk.Frame(parent)
        f.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        tk.Label(f, text="Download Measurement Template",
                 font=("TkDefaultFont", 12, "bold")).pack(anchor="w", pady=(0, 8))
        tk.Label(
            f,
            text=(
                "The template is pre-filled with your selected devices and their predicted values.\n"
                "Fill in the Meas Magnitude and Meas Angle columns in the field, "
                "then upload in the next step."
            ),
            wraplength=500, justify="left", fg="#444",
        ).pack(anchor="w", pady=(0, 16))

        tk.Button(f, text="Download Excel Template (.xlsx)",
                  font=("TkDefaultFont", 10), padx=8,
                  command=self._download_template).pack(anchor="w", pady=4)

        tk.Label(f, text="You can continue without downloading and upload later.",
                 fg="#888", font=("TkDefaultFont", 8)).pack(anchor="w", pady=(14, 0))

    def _download_template(self) -> None:
        import math
        from copy import copy
        from datetime import datetime
        import openpyxl

        TEMPLATE_PATH = Path(__file__).parent.parent.parent / "load_test_template.xlsx"
        if not TEMPLATE_PATH.exists():
            messagebox.showerror(
                "Template not found",
                f"Cannot find load_test_template.xlsx at:\n{TEMPLATE_PATH}",
                parent=self,
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*")],
            initialfile=f"LoadTest_{self._wo_var.get() or 'template'}.xlsx",
            title="Save Measurement Template",
            parent=self,
        )
        if not path:
            return

        # ── Template constants (match load_test_template.xlsx layout) ────────
        _STRIDE         = 6
        _FIRST_ROW      = 20
        _TEMPLATE_BLOCKS = 5
        _PHASE_COLS = {
            "A": ("U",  "U",  "Y"),
            "B": ("AD", "AD", "AH"),
            "C": ("AM", "AM", "AQ"),
            "N": ("AV", "AV", "AZ"),
        }

        def _stype(dtype: str) -> str:
            t = (dtype or "").lower()
            if t == "ct":        return "CT"
            if t == "vt":        return "VT"
            if t == "relay":     return "Relay"
            if t == "cttb":      return "CTTB"
            if t == "testblock": return "FT"
            return dtype.upper() if dtype else ""

        def _snap(c):
            return {"value": c.value, "font": copy(c.font), "border": copy(c.border),
                    "fill": copy(c.fill), "alignment": copy(c.alignment),
                    "number_format": c.number_format, "protection": copy(c.protection)}

        def _restore(c, s):
            c.value = s["value"]; c.font = s["font"]; c.border = s["border"]
            c.fill = s["fill"]; c.alignment = s["alignment"]
            c.number_format = s["number_format"]; c.protection = s["protection"]

        def _insert_rows(ws, idx, amount):
            affected = [rng for rng in list(ws.merged_cells.ranges) if rng.min_row >= idx]
            for rng in affected:
                ws.unmerge_cells(str(rng))
            ws.insert_rows(idx, amount)
            for rng in affected:
                ws.merge_cells(start_row=rng.min_row + amount, end_row=rng.max_row + amount,
                               start_column=rng.min_col, end_column=rng.max_col)

        def _delete_rows(ws, idx, amount):
            end = idx + amount
            for rng in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(rng))
                if rng.max_row < idx:
                    nmin, nmax = rng.min_row, rng.max_row
                elif rng.min_row >= end:
                    nmin, nmax = rng.min_row - amount, rng.max_row - amount
                elif rng.min_row < idx <= rng.max_row < end:
                    nmin, nmax = rng.min_row, idx - 1
                elif idx <= rng.min_row < rng.max_row < end:
                    continue
                else:
                    nmin, nmax = min(rng.min_row, idx), rng.max_row - amount
                if nmax >= nmin:
                    ws.merge_cells(start_row=nmin, end_row=nmax,
                                   start_column=rng.min_col, end_column=rng.max_col)
            ws.delete_rows(idx, amount)

        def _ensure_block_count(ws, n):
            n = max(n, 1)
            if n == _TEMPLATE_BLOCKS:
                return
            if n < _TEMPLATE_BLOCKS:
                _delete_rows(ws, _FIRST_ROW + n * _STRIDE,
                             _STRIDE * (_TEMPLATE_BLOCKS - n))
                return
            extra = n - _TEMPLATE_BLOCKS
            src = _FIRST_ROW + (_TEMPLATE_BLOCKS - 1) * _STRIDE
            ins = src + _STRIDE
            snap = [
                [_snap(ws.cell(row=src + dr, column=c))
                 for c in range(1, ws.max_column + 1)]
                for dr in range(_STRIDE)
            ]
            src_merges = [
                (rng.min_row - src, rng.max_row - src, rng.min_col, rng.max_col)
                for rng in ws.merged_cells.ranges
                if src <= rng.min_row < src + _STRIDE
            ]
            _insert_rows(ws, ins, _STRIDE * extra)
            for blk in range(extra):
                dst = ins + blk * _STRIDE
                for dr in range(_STRIDE):
                    for ci, s in enumerate(snap[dr], 1):
                        _restore(ws.cell(row=dst + dr, column=ci), s)
                for r0, r1, c0, c1 in src_merges:
                    ws.merge_cells(start_row=dst + r0, end_row=dst + r1,
                                   start_column=c0, end_column=c1)

        # ── Load & fill ───────────────────────────────────────────────────────
        from openpyxl.styles import PatternFill

        _TYPE_FILL = {
            "relay":     PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid"),
            "cttb":      PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid"),
            "testblock": PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid"),
            "ct":        PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid"),
            "vt":        PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid"),
        }

        wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
        if "Template" in wb.sheetnames:
            del wb["Template"]
        ws = wb["Load Test"]

        # Header
        ws["J4"]  = self._wo_var.get()
        ws["AM4"] = self._project_name
        ws["AX4"] = datetime.now().strftime("%Y-%m-%d")
        ws["AM5"] = self._tech_var.get()

        # Equipment list
        all_devs = self._all_devices()
        lines = ["Equipment:"]
        for pt in self._points:
            dev  = all_devs.get(pt.device_id)
            name = getattr(dev, "name", pt.device_name) if dev else pt.device_name
            lines.append(f"  - {name} ({pt.device_type.upper()})")
        ws["A8"] = "\n".join(lines)

        _ensure_block_count(ws, len(self._points))

        for i, pt in enumerate(self._points):
            r    = _FIRST_ROW + i * _STRIDE
            dev  = all_devs.get(pt.device_id)
            name = getattr(dev, "name", pt.device_name) if dev else pt.device_name
            stype = _stype(pt.device_type)
            ws.cell(row=r, column=1).value = stype
            ws[f"G{r}"]   = name
            ws[f"G{r+1}"] = getattr(dev, "location", "") or ""
            ws[f"C{r+2}"] = ""
            ws[f"L{r+2}"] = getattr(dev, "ratio_str", "") or getattr(dev, "ratio", "") or ""
            ws[f"J{r+3}"] = ""

            # Block row coloring by device type (matches old excel_report.py)
            fill = _TYPE_FILL.get(pt.device_type.lower())
            if fill:
                for dr in range(_STRIDE):
                    for dc in range(1, 53):
                        ws.cell(row=r + dr, column=dc).fill = fill

            # Phase labels in the label column of each phase (row r)
            # Phase column layout per block: label/mag at {col}{r+1} meas, {col}{r+2} predicted
            for phase, (label_col, mag_col, ang_col) in _PHASE_COLS.items():
                ws[f"{label_col}{r}"] = phase if phase != "N" else "Neutral"

            # Predicted values: balanced 3-phase from the single simulation value
            mag = pt.pred_magnitude
            ang = pt.pred_angle
            phase_data = {
                "A": (mag,   ang),
                "B": (mag,  (ang - 120) % 360),
                "C": (mag,  (ang + 120) % 360),
                "N": (0.0,   0.0),
            }
            for phase, (pmag, pang) in phase_data.items():
                _, mag_col, ang_col = _PHASE_COLS[phase]
                # Row r+1 = measured (blank — user fills in field)
                # Row r+2 = predicted secondary (pre-filled from simulation)
                c = ws[f"{mag_col}{r+2}"]
                c.value = round(pmag, 4)
                c.number_format = "0.0000"
                c = ws[f"{ang_col}{r+2}"]
                c.value = round(pang, 2)
                c.number_format = "0.00"

        try:
            wb.save(path)
            messagebox.showinfo("Downloaded", f"Template saved:\n{path}", parent=self)
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc), parent=self)

    # ── Step 7 (Manual): Measurement Order ───────────────────────────────────

    def _build_step_order(self, parent: tk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Measurement Order",
                 font=("TkDefaultFont", 12, "bold")).grid(
            row=0, column=0, sticky="w", padx=12, pady=(10, 2))
        tk.Label(parent,
                 text="Reorder using ↑ ↓. This sets the sequence in the measurement grid.",
                 wraplength=400, fg="#555").grid(
            row=0, column=1, sticky="w", padx=4, pady=(10, 2))

        self._order_listbox = tk.Listbox(parent, selectmode=tk.SINGLE, activestyle="dotbox")
        self._order_listbox.grid(row=1, column=0, sticky="nsew", padx=(12, 4), pady=(4, 10))

        btn_col = tk.Frame(parent)
        btn_col.grid(row=1, column=1, padx=(0, 12), pady=(4, 10), sticky="ns")
        tk.Button(btn_col, text="↑  Up",   command=self._order_up,   width=10).pack(pady=6)
        tk.Button(btn_col, text="↓  Down", command=self._order_down, width=10).pack(pady=6)

    def _refresh_order_list(self) -> None:
        all_devs = self._all_devices()
        lb = self._order_listbox
        lb.delete(0, tk.END)
        for pt in self._points:
            dev  = all_devs.get(pt.device_id)
            name = getattr(dev, "name", pt.device_name) if dev else pt.device_name
            lb.insert(tk.END, name)

    def _order_up(self) -> None:
        sel = self._order_listbox.curselection()
        if not sel or sel[0] == 0:
            return
        i = sel[0]
        self._points[i], self._points[i - 1] = self._points[i - 1], self._points[i]
        self._refresh_order_list()
        self._order_listbox.selection_set(i - 1)

    def _order_down(self) -> None:
        sel = self._order_listbox.curselection()
        if not sel or sel[0] >= len(self._points) - 1:
            return
        i = sel[0]
        self._points[i], self._points[i + 1] = self._points[i + 1], self._points[i]
        self._refresh_order_list()
        self._order_listbox.selection_set(i + 1)

    # ── Step 8 (Excel): Upload ────────────────────────────────────────────────

    def _build_step_upload(self, parent: tk.Frame) -> None:
        f = tk.Frame(parent)
        f.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        tk.Label(f, text="Upload Completed Measurements",
                 font=("TkDefaultFont", 12, "bold")).pack(anchor="w", pady=(0, 8))
        tk.Label(
            f,
            text=(
                "Select the completed Excel file to import measurements.\n"
                "Columns E and F must contain Meas Magnitude and Meas Angle."
            ),
            wraplength=500, justify="left", fg="#444",
        ).pack(anchor="w", pady=(0, 18))

        file_row = tk.Frame(f)
        file_row.pack(anchor="w", pady=4)
        self._upload_path_var = tk.StringVar(value="No file selected")
        tk.Label(file_row, textvariable=self._upload_path_var,
                 fg="#555", width=46, anchor="w", relief="sunken").pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(file_row, text="Browse…", command=self._upload_browse).pack(side=tk.LEFT)

        tk.Button(f, text="Import Measurements",
                  font=("TkDefaultFont", 10),
                  command=self._upload_import).pack(anchor="w", pady=(14, 4))
        tk.Button(f, text="Skip — Upload Later",
                  fg="#888",
                  command=self._upload_skip).pack(anchor="w")

    def _upload_browse(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*")],
            title="Select Completed Measurement File",
            parent=self,
        )
        if path:
            self._wiz_upload_path = path
            self._upload_path_var.set(path)

    def _upload_import(self) -> None:
        if not self._wiz_upload_path:
            messagebox.showwarning("No file", "Browse for a file first.", parent=self)
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._wiz_upload_path, data_only=True)
            ws = wb.active
            for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if ri >= len(self._points):
                    break
                try:
                    self._points[ri].meas_magnitude = float(row[4] or 0)
                    self._points[ri].meas_angle     = float(row[5] or 0) % 360
                except (TypeError, ValueError):
                    pass
            messagebox.showinfo("Imported", "Measurements imported.", parent=self)
        except Exception as exc:
            messagebox.showerror("Import failed", str(exc), parent=self)

    def _upload_skip(self) -> None:
        route = self._active_route()
        if 8 in route:
            pos = route.index(8)
            if pos < len(route) - 1:
                self._wiz_show(route[pos + 1])

    # ── Step 9: Measurements ──────────────────────────────────────────────────

    def _build_step_measurements(self, parent: tk.Frame) -> None:
        self._meas_nb = ttk.Notebook(parent)
        self._meas_nb.pack(fill=tk.BOTH, expand=True)

        meas_frame = tk.Frame(self._meas_nb)
        self._meas_nb.add(meas_frame, text="Measurements")
        self._build_measurements_content(meas_frame)

        vec_frame = tk.Frame(self._meas_nb)
        self._meas_nb.add(vec_frame, text="Vector Analysis")
        self._build_vector_content(vec_frame)

        diff_frame = tk.Frame(self._meas_nb)
        self._meas_nb.add(diff_frame, text="Differential")
        self._build_differential_content(diff_frame)

    def _build_measurements_content(self, parent: tk.Frame) -> None:
        self._meas_tab_frame = parent

        self._lock_label = tk.Label(
            parent,
            text="⛔  Measurements locked — confirm protection blocking in the Blocking step",
            font=("TkDefaultFont", 14, "bold"),
            fg="white", bg="red", wraplength=500,
        )

        self._meas_outer = tk.Frame(parent)
        self._meas_canvas = tk.Canvas(self._meas_outer)
        vsb = ttk.Scrollbar(self._meas_outer, orient="vertical",
                             command=self._meas_canvas.yview)
        hsb = ttk.Scrollbar(self._meas_outer, orient="horizontal",
                             command=self._meas_canvas.xview)
        self._meas_canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self._meas_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._meas_grid_frame = tk.Frame(self._meas_canvas)
        self._meas_win_id = self._meas_canvas.create_window(
            (0, 0), window=self._meas_grid_frame, anchor="nw")

        def _on_cfg(_event):
            self._meas_canvas.configure(scrollregion=self._meas_canvas.bbox("all"))

        self._meas_grid_frame.bind("<Configure>", _on_cfg)

        btn_bar = tk.Frame(parent)
        tk.Button(btn_bar, text="Recalculate", command=self._recalculate).pack(
            side=tk.LEFT, padx=4, pady=4)
        tk.Button(btn_bar, text="Save Test", command=self._save_test).pack(
            side=tk.LEFT, padx=4, pady=4)
        self._meas_btn_bar = btn_bar

    def _build_vector_content(self, parent: tk.Frame) -> None:
        canvas = tk.Canvas(parent)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inner = tk.Frame(canvas)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win_id, width=canvas.winfo_width())

        inner.bind("<Configure>", _cfg)
        canvas.bind("<Configure>", _cfg)
        pad = {"padx": 12, "pady": 6}

        vg = tk.LabelFrame(inner, text="3I₀ / V₀ Calculator", padx=8, pady=6)
        vg.pack(fill=tk.X, **pad)
        for c, h in enumerate(["Phase", "Magnitude", "Angle (°)"]):
            tk.Label(vg, text=h, font=("TkDefaultFont", 9, "bold")).grid(
                row=0, column=c, padx=6, pady=2)
        for r, (lbl, mv, av) in enumerate([
            ("A",              self._vec_a_mag, self._vec_a_ang),
            ("B",              self._vec_b_mag, self._vec_b_ang),
            ("C",              self._vec_c_mag, self._vec_c_ang),
            ("Neutral (meas.)",self._vec_n_mag, self._vec_n_ang),
        ], 1):
            tk.Label(vg, text=lbl, width=14, anchor="e").grid(row=r, column=0, padx=6, pady=2)
            tk.Entry(vg, textvariable=mv, width=12).grid(row=r, column=1, padx=6, pady=2)
            tk.Entry(vg, textvariable=av, width=12).grid(row=r, column=2, padx=6, pady=2)
        tk.Button(vg, text="Calculate", command=self._calc_vector).grid(
            row=5, column=0, columnspan=3, pady=6)
        self._vg_result_frame = tk.Frame(vg, relief="sunken", bd=1)
        self._vg_result_frame.grid(row=6, column=0, columnspan=3, sticky="ew", padx=4, pady=4)
        self._vg_result_lbl = tk.Label(
            self._vg_result_frame, textvariable=self._vec_result_var,
            font=("TkDefaultFont", 10), anchor="w", justify="left", padx=8,
        )
        self._vg_result_lbl.pack(anchor="w", fill=tk.X)

        fn = tk.LabelFrame(inner, text="Forcing Neutral Test", padx=8, pady=6)
        fn.pack(fill=tk.X, **pad)
        tk.Label(fn,
                 text="Short/isolate B & C phases. Measure A-phase and Neutral simultaneously.",
                 wraplength=520, justify="left").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 6))
        for c, h in enumerate(["", "Magnitude", "Angle (°)"]):
            tk.Label(fn, text=h, font=("TkDefaultFont", 9, "bold")).grid(
                row=1, column=c, padx=6, pady=2)
        for r, (lbl, mv, av) in enumerate([
            ("A-Phase", self._fn_a_mag, self._fn_a_ang),
            ("Neutral", self._fn_n_mag, self._fn_n_ang),
        ], 2):
            tk.Label(fn, text=lbl, width=14, anchor="e").grid(row=r, column=0, padx=6, pady=2)
            tk.Entry(fn, textvariable=mv, width=12).grid(row=r, column=1, padx=6, pady=2)
            tk.Entry(fn, textvariable=av, width=12).grid(row=r, column=2, padx=6, pady=2)
        tk.Button(fn, text="Check", command=self._check_forcing_neutral).grid(
            row=4, column=0, columnspan=3, pady=6)
        self._fn_result_lbl = tk.Label(
            fn, textvariable=self._fn_result_var,
            font=("TkDefaultFont", 10), anchor="w", justify="left", padx=8,
            relief="sunken", bd=1,
        )
        self._fn_result_lbl.grid(row=5, column=0, columnspan=3, sticky="ew", padx=4, pady=4)

    def _build_differential_content(self, parent: tk.Frame) -> None:
        self._diff_tab_frame = parent
        top = tk.Frame(parent)
        top.pack(fill=tk.X, padx=10, pady=6)
        tk.Label(top, text="CTTB Devices:",
                 font=("TkDefaultFont", 9, "bold")).pack(side=tk.LEFT)
        self._diff_listbox = tk.Listbox(parent, height=5, selectmode=tk.BROWSE)
        self._diff_listbox.pack(fill=tk.X, padx=10, pady=(0, 4))
        wiz = tk.LabelFrame(parent, text="Differential Wizard", padx=8, pady=6)
        wiz.pack(fill=tk.X, padx=10, pady=4)
        self._diff_step_lbl = tk.Label(
            wiz, text="Load equipment from SLD to begin.",
            font=("TkDefaultFont", 10), wraplength=480,
        )
        self._diff_step_lbl.pack(anchor="w", pady=4)
        entry_row = tk.Frame(wiz)
        entry_row.pack(anchor="w", pady=2)
        tk.Label(entry_row, text="Relay Reading — Magnitude:").pack(side=tk.LEFT)
        tk.Entry(entry_row, textvariable=self._diff_mag_var, width=10).pack(side=tk.LEFT, padx=4)
        tk.Label(entry_row, text="Angle (°):").pack(side=tk.LEFT)
        tk.Entry(entry_row, textvariable=self._diff_ang_var, width=10).pack(side=tk.LEFT, padx=4)
        nav_row = tk.Frame(wiz)
        nav_row.pack(anchor="w", pady=4)
        tk.Button(nav_row, text="◀ Previous", command=self._diff_prev).pack(side=tk.LEFT, padx=4)
        tk.Button(nav_row, text="Next ▶",     command=self._diff_next).pack(side=tk.LEFT, padx=4)
        self._diff_result_lbl = tk.Label(
            parent, text="", font=("TkDefaultFont", 11, "bold"), anchor="w",
        )
        self._diff_result_lbl.pack(fill=tk.X, padx=10, pady=4)
        self._diff_refresh_ui()

    # ── History tab ───────────────────────────────────────────────────────────

    def _build_history_tab(self, parent: tk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)

        cols = ("timestamp", "wo_number", "technologist", "device_count", "notes")
        self._hist_tree = ttk.Treeview(parent, columns=cols, show="headings",
                                        selectmode="browse")
        self._hist_tree.heading("timestamp",    text="Date / Time")
        self._hist_tree.heading("wo_number",    text="WO #")
        self._hist_tree.heading("technologist", text="Technologist")
        self._hist_tree.heading("device_count", text="Devices")
        self._hist_tree.heading("notes",        text="Notes")
        self._hist_tree.column("timestamp",    width=170, minwidth=150, stretch=False)
        self._hist_tree.column("wo_number",    width=120, minwidth=80,  stretch=False)
        self._hist_tree.column("technologist", width=160, minwidth=120, stretch=False)
        self._hist_tree.column("device_count", width=70,  minwidth=50,  stretch=False)
        self._hist_tree.column("notes",        width=200, minwidth=100, stretch=True)

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=self._hist_tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self._hist_tree.xview)
        self._hist_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._hist_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        btn_row = tk.Frame(parent)
        btn_row.grid(row=2, column=0, columnspan=2, sticky="ew", padx=6, pady=4)
        tk.Button(btn_row, text="Save Current Test",  command=self._save_test).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_row, text="Load Selected",      command=self._load_selected_test).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_row, text="Export CSV",         command=self._export_csv).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_row, text="Refresh",            command=self._refresh_history).pack(side=tk.RIGHT, padx=4)

    def _on_outer_tab_changed(self, _event) -> None:
        if self._outer_nb.index(self._outer_nb.select()) == 0:
            self._refresh_history()

    # ── Measurement grid logic ────────────────────────────────────────────────

    def _refresh_measurements_tab(self) -> None:
        # Excel path has no blocking step; unlock unconditionally
        locked = self._wiz_manual and not self._blocked.get()

        if locked:
            self._meas_outer.pack_forget()
            self._meas_btn_bar.pack_forget()
            self._lock_label.place(relx=0.5, rely=0.5, anchor="center")
            return

        self._lock_label.place_forget()
        self._meas_btn_bar.pack(side=tk.BOTTOM, fill=tk.X)
        self._meas_outer.pack(fill=tk.BOTH, expand=True)
        self._rebuild_grid()

    def _rebuild_grid(self) -> None:
        for w in self._meas_grid_frame.winfo_children():
            w.destroy()
        self._meas_vars.clear()

        headers    = ["Device", "Type", "Pred Mag", "Pred Ang°",
                      "Meas Mag", "Meas Ang°", "Δ Mag", "Δ Ang°", "Status"]
        col_widths = [10, 8, 9, 9, 9, 9, 8, 8, 7]

        for col, (hdr, w) in enumerate(zip(headers, col_widths)):
            tk.Label(
                self._meas_grid_frame, text=hdr,
                font=("TkDefaultFont", 9, "bold"),
                relief="ridge", width=w, bg="#D0D0D0",
            ).grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

        for row_idx, pt in enumerate(self._points):
            r  = row_idx + 1
            bg = "#FFFFFF" if row_idx % 2 == 0 else "#F5F5F5"

            def _lbl(text, col, bg=bg):
                tk.Label(self._meas_grid_frame, text=text, relief="flat",
                         bg=bg, anchor="w", padx=4).grid(
                    row=r, column=col, sticky="nsew", padx=1, pady=1)

            _lbl(pt.device_name,       0)
            _lbl(pt.device_type.upper(), 1)
            _lbl(f"{pt.pred_magnitude:.2f}", 2)
            _lbl(f"{pt.pred_angle:.1f}",     3)

            mag_var = tk.StringVar(value=f"{pt.meas_magnitude:.2f}")
            ang_var = tk.StringVar(value=f"{pt.meas_angle:.1f}")
            self._meas_vars[row_idx] = {"mag": mag_var, "ang": ang_var}

            tk.Entry(self._meas_grid_frame, textvariable=mag_var, width=9).grid(
                row=r, column=4, sticky="nsew", padx=1, pady=1)
            tk.Entry(self._meas_grid_frame, textvariable=ang_var, width=9).grid(
                row=r, column=5, sticky="nsew", padx=1, pady=1)

            tk.Label(self._meas_grid_frame, text=f"{pt.magnitude_delta:.4f}",
                     relief="flat", bg=bg, anchor="center").grid(
                row=r, column=6, sticky="nsew", padx=1, pady=1)
            tk.Label(self._meas_grid_frame, text=f"{pt.angle_delta:.2f}",
                     relief="flat", bg=bg, anchor="center").grid(
                row=r, column=7, sticky="nsew", padx=1, pady=1)

            status_text, status_fg = ("⚠ FLAG", "red") if pt.flagged else ("✓ OK", "green")
            tk.Label(self._meas_grid_frame, text=status_text,
                     fg=status_fg, font=("TkDefaultFont", 9, "bold"),
                     relief="flat", bg=bg).grid(
                row=r, column=8, sticky="nsew", padx=1, pady=1)

    def _recalculate(self) -> None:
        for idx, pt in enumerate(self._points):
            if idx in self._meas_vars:
                try:
                    pt.meas_magnitude = float(self._meas_vars[idx]["mag"].get())
                except ValueError:
                    pass
                try:
                    pt.meas_angle = float(self._meas_vars[idx]["ang"].get()) % 360
                except ValueError:
                    pass
        self._rebuild_grid()

    # ── Vector analysis ───────────────────────────────────────────────────────

    def _calc_vector(self) -> None:
        try:
            vg = VectorGroup(
                a_mag=float(self._vec_a_mag.get()), a_ang=float(self._vec_a_ang.get()),
                b_mag=float(self._vec_b_mag.get()), b_ang=float(self._vec_b_ang.get()),
                c_mag=float(self._vec_c_mag.get()), c_ang=float(self._vec_c_ang.get()),
                n_mag=float(self._vec_n_mag.get()), n_ang=float(self._vec_n_ang.get()),
            )
        except ValueError:
            messagebox.showerror("Input error", "All fields must be numeric.", parent=self)
            return
        import math
        calc_mag, calc_ang = vg.calculated_neutral()
        def ang_diff(a, b):
            d = (a - b) % 360
            return d if d <= 180 else d - 360
        mag_delta = round(calc_mag - vg.n_mag, 4)
        ang_delta = round(ang_diff(calc_ang, vg.n_ang), 2)
        pass_mag  = abs(mag_delta) <= calc_mag * 0.05 if calc_mag > 0 else abs(mag_delta) < 0.01
        pass_ang  = abs(ang_delta) <= 5.0
        verdict   = "PASS" if (pass_mag and pass_ang) else "FAIL"
        self._vec_result_var.set(
            f"Calculated Neutral:  {calc_mag:.4f} ∠ {calc_ang:.1f}°\n"
            f"Measured Neutral:    {vg.n_mag:.4f} ∠ {vg.n_ang:.1f}°\n"
            f"Δ Magnitude: {mag_delta:+.4f}   Δ Angle: {ang_delta:+.2f}°\n"
            f"Result: {verdict}"
        )
        self._vg_result_lbl.config(fg="green" if verdict == "PASS" else "red")

    def _check_forcing_neutral(self) -> None:
        try:
            a_pt = MeasurementPoint(
                device_id="A", device_name="A-Phase", device_type="ct",
                meas_magnitude=float(self._fn_a_mag.get()),
                meas_angle=float(self._fn_a_ang.get()),
            )
            n_pt = MeasurementPoint(
                device_id="N", device_name="Neutral", device_type="ct",
                meas_magnitude=float(self._fn_n_mag.get()),
                meas_angle=float(self._fn_n_ang.get()),
            )
        except ValueError:
            messagebox.showerror("Input error", "All fields must be numeric.", parent=self)
            return
        res     = check_forcing_neutral(a_pt, n_pt)
        overall = "PASS" if (res["mag_ok"] and res["ang_ok"]) else "FAIL"
        self._fn_result_var.set(
            f"Magnitude difference: {res['mag_diff']:.4f}  "
            f"[{'PASS' if res['mag_ok'] else 'FAIL'}]\n"
            f"Angle diff from 180°: {abs(abs(res['ang_diff'])-180):.2f}°  "
            f"[{'PASS' if res['ang_ok'] else 'FAIL'}]\n"
            f"Overall: {overall}"
        )
        self._fn_result_lbl.config(fg="green" if overall == "PASS" else "red")

    # ── Differential wizard ───────────────────────────────────────────────────

    def _diff_refresh_ui(self) -> None:
        n    = len(self._diff_cttbs)
        step = self._diff_step
        if n == 0:
            self._diff_step_lbl.config(
                text="No CTTB devices found. Load equipment from SLD in the Setup tab.")
            return
        if step >= n:
            self._diff_step_lbl.config(
                text=f"All {n} steps complete. Click 'Next' to calculate residual.")
            return
        self._diff_step_lbl.config(
            text=f"Step {step + 1} of {n} — Short/isolate all CTTBs except [{self._diff_cttbs[step]}].\n"
                 f"Enter the relay differential reading:")
        if step < len(self._diff_points):
            pt = self._diff_points[step]
            self._diff_mag_var.set(f"{pt.meas_magnitude:.2f}")
            self._diff_ang_var.set(f"{pt.meas_angle:.1f}")

    def _diff_prev(self) -> None:
        if self._diff_step > 0:
            self._diff_step -= 1
            self._diff_result_lbl.config(text="")
            self._diff_refresh_ui()

    def _diff_next(self) -> None:
        n = len(self._diff_cttbs)
        if n == 0:
            return
        if self._diff_step < n:
            try:
                mag = float(self._diff_mag_var.get())
                ang = float(self._diff_ang_var.get()) % 360
            except ValueError:
                messagebox.showerror("Input error",
                                     "Magnitude and Angle must be numeric.", parent=self)
                return
            pt = MeasurementPoint(
                device_id=self._diff_cttbs[self._diff_step],
                device_name=self._diff_cttbs[self._diff_step],
                device_type="cttb",
                meas_magnitude=mag, meas_angle=ang,
            )
            if self._diff_step < len(self._diff_points):
                self._diff_points[self._diff_step] = pt
            else:
                self._diff_points.append(pt)
            self._diff_step += 1
            self._diff_mag_var.set("0.0")
            self._diff_ang_var.set("0.0")
        if self._diff_step >= n:
            residual = differential_deviation(self._diff_points)
            verdict  = "PASS" if residual < 0.05 else "FAIL"
            self._diff_result_lbl.config(
                text=f"Differential residual: {residual:.4f} A  →  {verdict}",
                fg="green" if verdict == "PASS" else "red",
            )
        else:
            self._diff_result_lbl.config(text="")
            self._diff_refresh_ui()

    # ── SLD loading ───────────────────────────────────────────────────────────

    def _load_from_sld(self) -> None:
        if self._diagram is None:
            return
        d = self._diagram
        # Pre-select all devices from the SLD as a starting point
        self._wiz_selected_ids.clear()
        for attr in ("_cts", "_vts", "_cttbs", "_testblocks", "_relays"):
            self._wiz_selected_ids.extend(getattr(d, attr, {}).keys())

        # Differential CTTB list
        self._diff_cttbs = [
            getattr(dev, "name", dev_id)
            for dev_id, dev in getattr(d, "_cttbs", {}).items()
        ]
        self._diff_step = 0
        self._diff_points.clear()
        if hasattr(self, "_diff_listbox"):
            self._diff_listbox.delete(0, tk.END)
            for name in self._diff_cttbs:
                self._diff_listbox.insert(tk.END, name)
            self._diff_refresh_ui()

    def _on_blocking_changed(self) -> None:
        if hasattr(self, "_meas_tab_frame"):
            self._refresh_measurements_tab()

    # ── Persistence ───────────────────────────────────────────────────────────

    def _current_record(self) -> LoadTestRecord:
        for idx, pt in enumerate(self._points):
            if idx in self._meas_vars:
                try:
                    pt.meas_magnitude = float(self._meas_vars[idx]["mag"].get())
                except ValueError:
                    pass
                try:
                    pt.meas_angle = float(self._meas_vars[idx]["ang"].get()) % 360
                except ValueError:
                    pass
        return LoadTestRecord(
            id=str(uuid.uuid4()),
            project_name=self._project_name,
            wo_number=self._wo_var.get(),
            technologist=self._tech_var.get(),
            protection_blocked=self._blocked.get(),
            blocking_note=self._note_var.get(),
            points=[asdict(p) for p in self._points],
            vector_groups=[],
            drawings=list(self._wiz_test_drawings),
            notes="",
        )

    def _save_test(self) -> None:
        if not self._project_path:
            messagebox.showwarning(
                "No project file",
                "Save the project (File → Save) before saving a load test record.",
                parent=self,
            )
            return
        record = self._current_record()
        try:
            from poneglyph.io.project import save_load_test
            save_load_test(Path(self._project_path), asdict(record))
            self._refresh_history()
            messagebox.showinfo("Saved", f"Load test saved (id={record.id[:8]}…)", parent=self)
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc), parent=self)

    def _refresh_history(self) -> None:
        if not self._project_path:
            return
        try:
            from poneglyph.io.project import list_load_tests
            records = list_load_tests(Path(self._project_path))
        except Exception:
            return
        tree = self._hist_tree
        for item in tree.get_children():
            tree.delete(item)
        for rec in records:
            ts = rec.get("timestamp", "")[:19].replace("T", " ")
            tree.insert("", tk.END, iid=rec.get("id", ""), values=(
                ts,
                rec.get("wo_number", ""),
                rec.get("technologist", ""),
                len(rec.get("points", [])),
                rec.get("notes", ""),
            ))

    def _load_selected_test(self) -> None:
        sel = self._hist_tree.selection()
        if not sel:
            messagebox.showinfo("Nothing selected", "Select a record first.", parent=self)
            return
        try:
            from poneglyph.io.project import list_load_tests
            records = list_load_tests(Path(self._project_path))
        except Exception as exc:
            messagebox.showerror("Load failed", str(exc), parent=self)
            return
        rec = next((r for r in records if r.get("id") == sel[0]), None)
        if rec is None:
            messagebox.showerror("Not found", "Record not found.", parent=self)
            return
        self._wo_var.set(rec.get("wo_number", ""))
        self._tech_var.set(rec.get("technologist", ""))
        self._note_var.set(rec.get("blocking_note", ""))
        self._blocked.set(rec.get("protection_blocked", False))
        self._points.clear()
        valid_fields = set(MeasurementPoint.__dataclass_fields__)
        for p in rec.get("points", []):
            self._points.append(MeasurementPoint(**{k: v for k, v in p.items()
                                                    if k in valid_fields}))
        self._wiz_test_drawings = list(rec.get("drawings", []))
        self._meas_vars.clear()
        self._refresh_measurements_tab()
        messagebox.showinfo("Loaded", "Load test record restored.", parent=self)

    def _export_csv(self) -> None:
        sel = self._hist_tree.selection()
        if not sel:
            messagebox.showinfo("Nothing selected", "Select a record first.", parent=self)
            return
        try:
            from poneglyph.io.project import list_load_tests
            records = list_load_tests(Path(self._project_path))
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc), parent=self)
            return
        rec = next((r for r in records if r.get("id") == sel[0]), None)
        if rec is None:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*")],
            title="Export Load Test CSV",
            parent=self,
        )
        if not path:
            return
        try:
            with open(path, "w", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow([
                    "device_id", "device_name", "device_type", "channel",
                    "pred_magnitude", "pred_angle",
                    "meas_magnitude", "meas_angle",
                ])
                for p in rec.get("points", []):
                    writer.writerow([
                        p.get("device_id"), p.get("device_name"),
                        p.get("device_type"), p.get("channel"),
                        p.get("pred_magnitude"), p.get("pred_angle"),
                        p.get("meas_magnitude"), p.get("meas_angle"),
                    ])
            messagebox.showinfo("Exported", f"CSV saved to {path}", parent=self)
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc), parent=self)

    # ── Project update ────────────────────────────────────────────────────────

    def set_project(self, project_name: str = "", project_path=None) -> None:
        self._project_name = project_name
        if project_path is not None:
            self._project_path = project_path
        self._refresh_history()
